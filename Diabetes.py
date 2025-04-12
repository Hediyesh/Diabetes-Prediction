import math
import pandas as pd
from sklearn.mixture import GaussianMixture
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.metrics import jaccard_score, accuracy_score, precision_score, f1_score, recall_score
import numpy as np
import openpyxl as xl
import random
import seaborn as sns
from sklearn import metrics
import matplotlib.pyplot as plt
from openpyxl.workbook import Workbook
import warnings


#heatmap of data
def heatmap(df):
    plt.figure(figsize=(12, 8))
    sns.heatmap(df.corr(), annot=True, cmap="YlGnBu")
    plt.savefig('heatmap_plot.jpg', format='jpg')
    plt.show()


#scatter plot of data
def scatterPlot(data):
    dff = data.copy()
    sns.pairplot(data=dff, hue='Outcome', kind='scatter', palette='bright')
    plt.show()


#box plot of data
def boxPlot(df):
    num_col = ['Pregnancies', 'Glucose', 'BloodPressure', 'SkinThickness'
        , 'Insulin', 'BMI', 'DiabetesPedigreeFunction', 'Age', 'Outcome']
    no_outlier = df
    for i in num_col:
        lower_limit = df[i].quantile(0.25)
        upper_limit = df[i].quantile(0.75)
        no_outlier[i] = no_outlier[i].clip(lower_limit, upper_limit)
    df.plot(kind="box", subplots=True, figsize=(15, 15), layout=(5, 5))
    plt.show()


#pie chart of outcomes 0 and 1
def piechart(df):
    outcome_counts = df["Outcome"].value_counts()
    outcome_counts.plot(kind="pie", autopct="%1.1f%%", startangle=90)
    plt.title("Distribution of Outcomes")
    plt.ylabel("")
    plt.show()


#data feature description
def describe(df):
    summary = df.describe()
    fig, ax = plt.subplots(figsize=(8, 4))
    table_data = []
    for col in summary.columns:
        table_data.append([col] + list(summary[col].values))

    table = ax.table(cellText=table_data,
                     colLabels=['Statistic'] + ['count', 'mean', 'std', 'min', '25%', '50%', '70%', 'max'],
                     cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.2)
    ax.axis('off')
    plt.show()


# separating diabetes data into data with null and data without null
def count_null():
    wb1 = xl.load_workbook('diabetes_700/diabetes_normal.xlsx')
    sheet1 = wb1['Sheet1']
    wb2 = xl.load_workbook('diabetes_700/diabetes_without_null.xlsx')
    sheet2 = wb2['Sheet1']
    wb3 = xl.load_workbook('diabetes_700/diabetes_with_null.xlsx')
    sheet3 = wb3['Sheet1']
    r2 = 2
    r3 = 2
    for j in range(2, sheet1.max_row+1):
        c = 0
        # if there is at least one invalid column, the row will be added to invalid data
        # otherwise will be added to valid data
        for i in range(2, 9):
            if sheet1.cell(j, i).value == -1:
                c += 1
        if c == 0:
            for k in range(1, 10):
                sheet2.cell(r2, k).value = sheet1.cell(j, k).value
            r2 += 1
        else:
            for k in range(1, 10):
                sheet3.cell(r3, k).value = sheet1.cell(j, k).value
            r3 += 1
    wb2.save('diabetes_700/diabetes_without_null.xlsx')
    wb3.save('diabetes_700/diabetes_with_null.xlsx')


# normalization
def normal_data():
    wb1 = xl.load_workbook('diabetes_700/diabetes.xlsx')
    sheet1 = wb1['Sheet1']
    wb2 = xl.load_workbook('diabetes_700/diabetes_normal.xlsx')
    sheet2 = wb2['Sheet1']
    for col in range(1, 9):
        max1 = 0
        min1 = 100
        # find min and max
        for row in range(2, sheet1.max_row+1):
            if float(sheet1.cell(row, col).value) > max1:
                max1 = float(sheet1.cell(row, col).value)
            if float(sheet1.cell(row, col).value) < min1 and float(sheet1.cell(row, col).value) != -1:
                min1 = float(sheet1.cell(row, col).value)
        # replace data with min and max normalization
        for row in range(2, sheet1.max_row+1):
            print(f'number:{sheet1.cell(row, col).value}')
            if sheet1.cell(row, col).value != -1:
                sheet2.cell(row, col).value = (float(sheet1.cell(row, col).value) - min1)/(max1 - min1)
                print(f'max:{max1},min:{min1},normal:{sheet2.cell(row, col).value}')
            else:
                sheet2.cell(row, col).value = -1
    wb2.save('diabetes_700/diabetes_normal.xlsx')


# 1st step of classification = counting 70% zero and one outcomes and 30% of them
def count_zeroandone(sheet_normal):
    zeroes = 0
    ones = 0
    # count zeroes and one outcomes and find 30% and 70% number, (row, 9).value == 0 for diabetes data, 9 is 1 for bc
    for row in range(2, sheet_normal.max_row+1):
        if sheet_normal.cell(row, 1).value == 0:
            zeroes += 1
        else:
            ones += 1
    test_zero = round(zeroes*30/100)
    test_one = round(ones*30/100)
    print(f'data has {sheet_normal.max_row-1} rows')
    print(f'data has {zeroes} zero and {ones} one outcome.')
    print(f'30% of zeroes: {test_zero}\n30% of ones: {test_one}')
    print(f'70% of zeroes: {zeroes-test_zero}\n70% of ones: {ones-test_one}')


# 2nd step of classification = splitting data into diabetic and non-diabetic data
def split_zeroandone(sheet_normal):
    # wb1 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/12/zeroes.xlsx')
    wb1 = xl.load_workbook('breast_cancer/zeroes.xlsx')
    sheet_zero = wb1['Sheet1']
    # wb2 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/12/ones.xlsx')
    wb2 = xl.load_workbook('breast_cancer/ones.xlsx')
    sheet_one = wb2['Sheet1']
    zeroc = 0
    onec = 0
    row1 = 2
    row2 = 2
    # add feature names to excels, range(1, 10) for diabetes, 10 is 32 for bc
    for names in range(1, 32):
        sheet_zero.cell(1, names).value = sheet_normal.cell(1, names).value
        sheet_one.cell(1, names).value = sheet_normal.cell(1, names).value
    # based on outcome, add data to zeroes and ones
    for row in range(2, sheet_normal.max_row+1):
        if sheet_normal.cell(row, 1).value == 0:
            zeroc += 1
            for col in range(1, 32):
                sheet_zero.cell(row1, col).value = sheet_normal.cell(row, col).value
            row1 += 1
        else:
            onec += 1
            for col in range(1, 32):
                sheet_one.cell(row2, col).value = sheet_normal.cell(row, col).value
            row2 += 1
    print(f'{zeroc} zeroes and {onec} ones were placed.')
    # wb2.save('diabetes_700/ways/thesis_ch4.2/12/ones.xlsx')
    # wb1.save('diabetes_700/ways/thesis_ch4.2/12/zeroes.xlsx')
    wb2.save('breast_cancer/ones.xlsx')
    wb1.save('breast_cancer/zeroes.xlsx')


# 3rd step of classification = splitting 30% test data and 70% train data
def split_train_test(sheet_normal):
    # wb1 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/59/test.xlsx')
    wb1 = xl.load_workbook('breast_cancer/2/test.xlsx')
    sheet_test = wb1['Sheet1']
    # wb2 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/59/zeroes.xlsx')
    wb2 = xl.load_workbook('breast_cancer/zeroes.xlsx')
    sheet_zero = wb2['Sheet1']
    # wb3 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/59/ones.xlsx')
    wb3 = xl.load_workbook('breast_cancer/ones.xlsx')
    sheet_one = wb3['Sheet1']
    # 30% of data range for outcome 0 and 1
    # range0 = 150 for diabetes
    # range1 = 80 for diabetes
    range0 = 107
    range1 = 64
    row = 2
    # add feature names to test data, range(1, 10) for d
    for names in range(1, 32):
        sheet_test.cell(1, names).value = sheet_normal.cell(1, names).value
    # add zeroes to test data
    for i in range(1, range0+1):
        rand_num = random.randint(2, sheet_zero.max_row+1)
        for col in range(1, 32):
            sheet_test.cell(row, col).value = sheet_zero.cell(rand_num, col).value
        sheet_zero.delete_rows(rand_num)
        # wb2.save('diabetes_700/ways/thesis_ch4.2/59/zeroes.xlsx')
        wb2.save('breast_cancer/zeroes.xlsx')
        row += 1
    # add ones to test data
    for i in range(1, range1+1):
        rand_num = random.randint(2, sheet_one.max_row+1)
        for col in range(1, 32):
            sheet_test.cell(row, col).value = sheet_one.cell(rand_num, col).value
        sheet_one.delete_rows(rand_num)
        # wb3.save('diabetes_700/ways/thesis_ch4.2/59/ones.xlsx')
        wb3.save('breast_cancer/ones.xlsx')
        row += 1
    # rest of ones and zeroes will make train data
    # wb2.save('diabetes_700/ways/thesis_ch4.2/59/zeroes.xlsx')
    # wb3.save('diabetes_700/ways/thesis_ch4.2/59/ones.xlsx')
    # wb1.save('diabetes_700/ways/thesis_ch4.2/59/test.xlsx')
    wb2.save('breast_cancer/zeroes.xlsx')
    wb3.save('breast_cancer/ones.xlsx')
    wb1.save('breast_cancer/2/test.xlsx')


# 4th step of classification = make excel when needed for k 2 to 10
def make_excel():
    wb2 = Workbook()
    for i in range(2, 11):
        # wb2.save("diabetes_700/ways/thesis_ch4.2/1/k"+str(i)+".xlsx")
        wb2.save("breast_cancer/1/k"+str(i)+".xlsx")


# 5th step of classification = GMM clustering on train data
def GMM():
    # data1 = pd.read_excel('diabetes_700/ways/thesis_ch4.2/59/train.xlsx')
    data1 = pd.read_excel('breast_cancer/2/train.xlsx')
    for k in range(2, 11):
        # selected = data1.drop(columns=['Outcome'])
        selected = data1.drop(columns=['diagnosis'])
        gmm = GaussianMixture(n_components=k)
        cluster_labels = gmm.fit_predict(selected)
        data1['Cluster'] = cluster_labels
        # data1.to_excel("diabetes_700/ways/thesis_ch4.2/59/k"+str(k)+".xlsx", index=False)
        # wb1 = xl.load_workbook("diabetes_700/ways/thesis_ch4.2/59/k"+str(k)+".xlsx")
        data1.to_excel("breast_cancer/2/k"+str(k)+".xlsx", index=False)
        wb1 = xl.load_workbook("breast_cancer/2/k"+str(k)+".xlsx")
        sheet1 = wb1['Sheet1']
        print(f'k={k}')
        # printing how many zeroe and one outcomes are there in clusters for each k
        # sheet1.cell(row, 10).value == cl, sheet1.cell(row, 9).value == 0, for d
        for cl in set(cluster_labels):
            c = 0
            out0 = 0
            out1 = 0
            for row in range(2, sheet1.max_row + 1):
                if sheet1.cell(row, 32).value == cl:
                    c += 1
                    if sheet1.cell(row, 1).value == 0:
                        out0 += 1
                    else:
                        out1 += 1
            print(f'cluster {cl} has count={c}. There are {out0} zeroes and {out1} ones.')


# 6th step of classification = label test data with diabetic and non-diabetic clusters using threshold
def method():
    # define non-diabetic and diabetic and test data
    # c1 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/.../....xlsx')
    # c2 = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/.../....xlsx')
    # test_data = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/.../....xlsx')
    c1 = xl.load_workbook('breast_cancer/2/k3_1_1.xlsx')
    c2 = xl.load_workbook('breast_cancer/2/k3_0_0.xlsx')
    test_data = xl.load_workbook('breast_cancer/2/test.xlsx')
    sheet_c1 = c1['Sheet1']
    sheet_c2 = c2['Sheet1']
    sheet_test = test_data['Sheet1']
    # define threshold, cin=count of data within threshold limit, cout=count of data outside threshold limit
    thrs = 0.6
    cin = 0
    cout = 0
    for row in range(2, sheet_test.max_row + 1):
        # calculate min distance of test data from 2 clusters, for g in range(1, 9) for d
        min_dis21 = 0
        for g in range(2, 32):
            min_dis21 += math.pow(float(sheet_c1.cell(2, g).value) - float(sheet_test.cell(row, g).value), 2)
        min_dis21 = math.sqrt(min_dis21)
        min_dis22 = 0
        for g in range(2, 32):
            min_dis22 += math.pow(float(sheet_c2.cell(2, g).value) - float(sheet_test.cell(row, g).value), 2)
        min_dis22 = math.sqrt(min_dis22)
        min_row1 = 2
        min_row2 = 2
        for row1 in range(2, sheet_c1.max_row + 1):
            min_dis1 = 0
            for g in range(2, 32):
                min_dis1 += math.pow(float(sheet_c1.cell(row1, g).value) - float(sheet_test.cell(row, g).value), 2)
            min_dis1 = math.sqrt(min_dis1)
            if min_dis1 < min_dis21:
                min_dis21 = min_dis1
                min_row1 = row1
        for row1 in range(2, sheet_c2.max_row + 1):
            min_dis2 = 0
            for g in range(2, 32):
                min_dis2 += math.pow(float(sheet_c2.cell(row1, g).value) - float(sheet_test.cell(row, g).value), 2)
            min_dis2 = math.sqrt(min_dis2)
            if min_dis2 < min_dis22:
                min_dis22 = min_dis2
                min_row2 = row1
        # find the minimum distance and if it is less than the threshold, test data will be labeled
        # sheet_test.cell(row, 10).value = sheet_c2.cell(min_row2, 10).value for d
        if min_dis22 < min_dis21:
            if min_dis22 < thrs:
                cin += 1
                sheet_test.cell(row, 32).value = sheet_c2.cell(min_row2, 32).value
            else:
                cout += 1
        if min_dis21 < min_dis22:
            if min_dis21 < thrs:
                sheet_test.cell(row, 32).value = sheet_c1.cell(min_row1, 32).value
                cin += 1
            else:
                cout += 1
    print(cin, cout)
    # test_data.save('diabetes_700/ways/thesis_ch4.2/1/test.xlsx')
    test_data.save('breast_cancer/2/test.xlsx')


# 7th step of classification = performing randomforest on rejected data
def randomforest():
    # define the impure cluster and rejected data
    # train_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/.../k3_rest.xlsx')
    # test_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/.../rejected.xlsx')
    # X_train = train_data.drop(columns=['Outcome'])  # features for rf
    # y_train = train_data['Outcome']  # target variable
    # X_test = test_data.drop(columns=['Outcome'])  # features for rf
    train_data = pd.read_excel('breast_cancer/2/k3_rest.xlsx')
    test_data = pd.read_excel('breast_cancer/2/rejected.xlsx')
    X_train = train_data.drop(columns=['diagnosis'])  # features for rf
    y_train = train_data['diagnosis']  # target variable
    X_test = test_data.drop(columns=['diagnosis'])  # features for rf
    # perform rf
    n_estimators = 50
    random_forest = RandomForestClassifier(max_depth=5, n_estimators=n_estimators, random_state=42)
    random_forest.fit(X_train, y_train)
    y_pred = random_forest.predict(X_test)
    # y_true = test_data['Outcome']
    y_true = test_data['diagnosis']
    train_prd = random_forest.predict(X_train)
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred)
    recall = recall_score(y_true, y_pred)
    f1 = f1_score(y_true, y_pred)
    # save data
    test_data['label'] = y_pred
    # test_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/rejected.xlsx', sheet_name='Sheet1', index=False)
    test_data.to_excel('breast_cancer/2/rejected.xlsx', sheet_name='Sheet1', index=False)
    train_data['label'] = train_prd
    # train_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/k3_rest.xlsx', sheet_name='Sheet1', index=False)
    # train_data.to_excel('breast_cancer/1/k3_rest.xlsx', sheet_name='Sheet1', index=False)
    # print the evaluation metrics
    print("Accuracy:", accuracy)
    print("Precision:", precision)
    print("Recall", recall)
    print("F1 Score:", f1)


# 8th step of classification = calculating evaluation metrics
def data_confusion_matrix():
    # predicted = np.array(pd.read_excel('diabetes_700/ways/thesis_ch4.2/.../test.xlsx')['label'])
    # actual = np.array(pd.read_excel('diabetes_700/ways/thesis_ch4.2/.../test.xlsx')['Outcome'])
    predicted = np.array(pd.read_excel('breast_cancer/2/test.xlsx')['label'])
    actual = np.array(pd.read_excel('breast_cancer/2/test.xlsx')['diagnosis'])
    confusion_matrix = metrics.confusion_matrix(actual, predicted)
    precision = confusion_matrix[1, 1] / (confusion_matrix[1, 1] + confusion_matrix[0, 1])
    recall = confusion_matrix[1, 1] / (confusion_matrix[1, 1] + confusion_matrix[1, 0])
    f1_score = 2 * (precision * recall) / (precision + recall)
    accuracy = (confusion_matrix[0, 0] + confusion_matrix[1, 1]) / np.sum(confusion_matrix)
    print("Recall:", recall)
    print("Accuracy:", accuracy)
    print("Precision:", precision)
    print("F1-Score:", f1_score)
    cm_display = metrics.ConfusionMatrixDisplay(confusion_matrix=confusion_matrix, display_labels=[False, True])
    cm_display.plot()
    plt.title('Confusion Matrix')
    plt.show()


# performing svm on test data
def svm():
    train_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/1/k3_rest.xlsx')
    test_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/1/rejected.xlsx')
    X_train = train_data.drop(columns=['Outcome', 'label'])
    y_train = train_data['Outcome']
    # 'sigmoid', 'linear', 'rbf', 'poly', poly was better
    # perform svm
    svm_classifier = SVC(kernel='poly')
    svm_classifier.fit(X_train, y_train)
    X_test = test_data.drop(columns=['Outcome'])  # Features
    y_pred_train = svm_classifier.predict(X_train)
    train_data['label'] = y_pred_train
    y_pred = svm_classifier.predict(X_test)
    y_test = test_data['Outcome']
    test_data['label'] = y_pred
    train_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/k3_rest.xlsx', index=False)
    test_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/rejected.xlsx', index=False)
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    recall = recall_score(y_test, y_pred)
    print("Accuracy:", accuracy)
    print("Precision:", precision)
    print("F1 Score:", f1)
    print("Recall", recall)


# performing decision tree on test data
def dtree():
    train_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/1/k3_rest.xlsx')
    test_data = pd.read_excel('diabetes_700/ways/thesis_ch4.2/1/rejected.xlsx')
    # extract features and target variable from train and test data
    X_train = train_data.drop(columns=['Outcome'])
    y_train = train_data['Outcome']
    X_test = test_data.drop(columns=['Outcome'])
    y_test = test_data['Outcome']
    # Train Decision Tree model
    decision_tree = DecisionTreeClassifier(max_depth=7, random_state=4)
    decision_tree.fit(X_train, y_train)
    y_pred_train = decision_tree.predict(X_train)
    y_pred_test = decision_tree.predict(X_test)
    train_data['label'] = y_pred_train
    test_data['label'] = y_pred_test
    # save train and test data
    train_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/k3_rest.xlsx', index=False)
    test_data.to_excel('diabetes_700/ways/thesis_ch4.2/1/rejected.xlsx', index=False)
    # calculate evaluation metrics for test data
    jaccard = jaccard_score(y_test, y_pred_test)
    accuracy = accuracy_score(y_test, y_pred_test)
    precision = precision_score(y_test, y_pred_test)
    f1 = f1_score(y_test, y_pred_test)
    print("Jaccard Similarity (Test Data):", jaccard)
    print("Accuracy (Test Data):", accuracy)
    print("Precision (Test Data):", precision)
    print("F1 Score (Test Data):", f1)


# define data for replacing invalid data with the proposed method
wb_tst = xl.load_workbook('diabetes_700/ways/diabetes_with_null.xlsx')
sheet_tst = wb_tst['Sheet1']
wb_trn = xl.load_workbook('diabetes_700/ways/wout_nulls_c/k2.xlsx')
sheet_trn = wb_trn['Sheet1']
wb_k0 = xl.load_workbook('diabetes_700/ways/wout_nulls_c/k2_0.xlsx')
sheet_k0 = wb_k0['Sheet1']
wb_k1 = xl.load_workbook('diabetes_700/ways/wout_nulls_c/k2_1.xlsx')
sheet_k1 = wb_k1['Sheet1']
data_k0 = pd.read_excel('diabetes_700/ways/wout_nulls_c/k2_0.xlsx')
center_c0 = data_k0.groupby('Cluster').mean().values.tolist()
data_k1 = pd.read_excel('diabetes_700/ways/wout_nulls_c/k2_1.xlsx')
center_c1 = data_k1.groupby('Cluster').mean().values.tolist()


# calculating distance from cluster center based on number of row and cluster
def dis_c(s, k):
    if k == 1:
        dis = 0
        for col in range(1, 9):
            dis += abs(float(sheet_k1.cell(s, col).value) - float(center_c0[0][col-1]))
        dis = math.sqrt(dis)
        return dis
    else:
        dis = 0
        for col in range(1, 9):
            dis += abs(float(sheet_k0.cell(s, col).value) - float(center_c1[0][col - 1]))
        dis = math.sqrt(dis)
        return dis


# the proposed method for missing value imputation
def fill_null_with_cluster():
    for row in range(2, sheet_tst.max_row+1):
        print(f'row:{row}')
        # check the value of outcome
        if sheet_tst.cell(row, 9).value == 1:
            li_sort = []
            # outcome is 1, so calculate data node distances from mixture cluster
            for row1 in range(2, sheet_k1.max_row+1):
                if sheet_k1.cell(row1, 9).value == 1:
                    li = []
                    li_main = []
                    li_main.append(sheet_tst.cell(row, 1).value)
                    li.append(sheet_k1.cell(row1, 1).value)
                    for c in range(2, 9):
                        if sheet_tst.cell(row, c).value != -1:
                            li.append(sheet_k1.cell(row1, c).value)
                            li_main.append(sheet_tst.cell(row, c).value)
                    distance = 0
                    for k in range(0, len(li)):
                        minus = abs(float(li[k]) - float(li_main[k])) ** 2
                        distance += minus
                    distance_sqrt = math.sqrt(distance)
                    li_sort.append([distance_sqrt, row1])
            sorted_list = sorted(li_sort, key=lambda x: x[0])
            # select 10 nearest
            selected = [sub[1] for sub in sorted_list][:10]
            # print(selected)
            max1 = 0
            rs = 0
            # calculate distance from non-diabetic cluster center and select furthest one
            for s in selected:
                dis = dis_c(s, 1)
                # print(dis, s)
                if dis > max1:
                    max1 = dis
                    rs = s
            print(f'1: {rs}')
            cout = 0
            # replace invalid data
            for col1 in range(1, 9):
                if sheet_tst.cell(row, col1).value == -1:
                    cout += 1
                    sheet_tst.cell(row, col1).value = sheet_k1.cell(rs, col1).value
            print(f'{cout} replaced')
        else:
            li_sort0 = []
            li_sort1 = []
            # outcome is zero, distance from both cluster nodes are calculated
            for row1 in range(2, sheet_k1.max_row + 1):
                if sheet_k1.cell(row1, 9).value == 0:
                    li = []
                    li_main = []
                    li_main.append(sheet_tst.cell(row, 1).value)
                    li.append(sheet_k1.cell(row1, 1).value)
                    for c in range(2, 9):
                        if sheet_tst.cell(row, c).value != -1:
                            li.append(sheet_k1.cell(row1, c).value)
                            li_main.append(sheet_tst.cell(row, c).value)
                    distance = 0
                    for k in range(0, len(li)):
                        minus = abs(float(li[k]) - float(li_main[k])) ** 2
                        distance += minus
                    distance_sqrt = math.sqrt(distance)
                    li_sort1.append([distance_sqrt, row1])
            # 10 nearest are selected from both
            sorted_list1 = sorted(li_sort1, key=lambda x: x[0])
            selected1 = [sub[0] for sub in sorted_list1][:1]
            for row1 in range(2, sheet_k0.max_row + 1):
                if sheet_k0.cell(row1, 9).value == 0:
                    li = []
                    li_main = []
                    li_main.append(sheet_tst.cell(row, 1).value)
                    li.append(sheet_k0.cell(row1, 1).value)
                    for c in range(2, 9):
                        if sheet_tst.cell(row, c).value != -1:
                            li.append(sheet_k0.cell(row1, c).value)
                            li_main.append(sheet_tst.cell(row, c).value)
                    distance = 0
                    for k in range(0, len(li)):
                        minus = abs(float(li[k]) - float(li_main[k])) ** 2
                        distance += minus
                    distance_sqrt = math.sqrt(distance)
                    li_sort0.append([distance_sqrt, row1])
            sorted_list0 = sorted(li_sort0, key=lambda x: x[0])
            selected0 = [sub[0] for sub in sorted_list0][:1]
            if selected0 < selected1:
                # if data node is closer to non-diabetic cluster, select 10 colsest
                selected00 = [sub[1] for sub in sorted_list0][:10]
                max1 = 0
                rs = 0
                # select closest one to non-diabetic cluster
                for s in selected00:
                    dis = dis_c(s, 0)
                    # print(dis, s)
                    if dis > max1:
                        max1 = dis
                        rs = s
                print(f'0 k0: {rs}')
                cout = 0
                # replace invalid data
                for col1 in range(1, 9):
                    if sheet_tst.cell(row, col1).value == -1:
                        cout += 1
                        sheet_tst.cell(row, col1).value = sheet_k0.cell(rs, col1).value
                print(f'{cout} replaced')
            else:
                # if data node is closer to mixture cluster, select 10 colsest
                selected11 = [sub[1] for sub in sorted_list1][:10]
                min1 = 100
                rs = 0
                # select furthest one from mixture cluster
                for s in selected11:
                    dis = dis_c(s, 1)
                    # print(dis, s)
                    if dis < min1:
                        min1 = dis
                        rs = s
                print(f'0 k1: {rs}')
                cout = 0
                # replace invalid data
                for col1 in range(1, 9):
                    if sheet_tst.cell(row, col1).value == -1:
                        cout += 1
                        sheet_tst.cell(row, col1).value = sheet_k1.cell(rs, col1).value
                print(f'{cout} replaced')
        print('_____________')
    # save data
    wb_tst.save('diabetes_700/ways/diabetes_with_null.xlsx')


# replace invalid data with mean of feature
def replace_mean_invalid():
    wb_mean = xl.load_workbook('diabetes_700/ways/datawithoutnull/replace_mean_invalid.xlsx')
    sheet_mean = wb_mean['Sheet1']
    # find average of valid data for features and replace invalid data with them
    for col in range(2, 7):
        sum1 = 0.0
        c = 0.0
        for row in range(2, sheet_mean.max_row):
            if sheet_mean.cell(row, col).value != -1:
                sum1 = sum1 + sheet_mean.cell(row, col).value
                c += 1
        avg = sum1/c
        print(col, avg)
        for row in range(2, sheet_mean.max_row):
            if sheet_mean.cell(row, col).value == -1:
                sheet_mean.cell(row, col).value = avg
        wb_mean.save('diabetes_700/ways/datawithoutnull/replace_mean_invalid.xlsx')


# replace invalid data with mean of features in the same class(diabetic or non-diabetic)
def mean_of_not_null_classes():
    wb_all = xl.load_workbook('diabetes_700/ways/ch5/alldiabetes_normal.xlsx')
    sheet_all = wb_all['Sheet1']
    # for each sample of data, find average of features based on outcome classes
    # and replace invalid data of the sample with the averages
    for row in range(2, sheet_all.max_row + 1):
        for i in range(2, 7):
            if sheet_all.cell(row, i).value == -1:
                sum = 0.0
                count = 0.0
                for row2 in range(2, sheet_all.max_row + 1):
                    if sheet_all.cell(row, 9).value == sheet_all.cell(row2, 9).value and sheet_all.cell(row2, i).value != -1:
                        sum += sheet_all.cell(row2, i).value
                        count += 1
                sheet_all.cell(row, i).value = sum/float(count)
                # print(f'row:{row}, col:{i}, class:{sheet_all.cell(row, 9).value}, val:{sheet_all.cell(row, i).value}')
    wb_all.save('diabetes_700/ways/ch5/alldiabetes_normal.xlsx')
    # making sure all data is valid(c would be 0)
    c = 0
    for row in range(2, sheet_all.max_row + 1):
        for i in range(2, 7):
            if sheet_all.cell(row, i).value == -1:
                c += 1
    print(c)

    # Load the data
def visualization():
    dt = pd.read_excel('diabetes_700/ways/way2/k3_train.xlsx')
    # Convert fraction strings to float
    for col in dt.columns:
        if dt[col].dtype == 'object':
            dt[col] = dt[col].str.replace('/', '').astype(float)
    # Set the style of the plot
    sns.set(style='whitegrid')
    # Define custom color mapping for each combination of Outcome and Cluster
    color_mapping = {
        (0, 0): 'green',
        (0, 1): 'blue',
        (0, 2): 'purple',
        (1, 0): 'red',
        (1, 1): 'pink',
        (1, 2): 'yellow'
    }
    # Plot the data
    plt.figure(figsize=(10, 6))
    for idx, row in dt.iterrows():
        outcome = row['Outcome']
        cluster = row['Cluster']
        color = color_mapping.get((outcome, cluster), 'gray')
        marker = 'o' if outcome == 0 else 'o'  # Use circles for outcome 0 and squares for outcome 1
        plt.scatter(row['Glucose'], row['BMI'], marker=marker, color=color)
    # Create custom legend
    legend_labels = {
        'Non-Diabetic Cluster 1': 'green',
        'Non-Diabetic Cluster 2': 'blue',
        'Non-Diabetic Cluster 3': 'purple',
        'Diabetic Cluster 1': 'red',
        'Diabetic Cluster 2': 'pink',
        'Diabetic Cluster 3': 'yellow'
    }
    # Plot invisible points for creating custom legend
    for label, color in legend_labels.items():
        plt.scatter([], [], color=color, label=label)
    # Add legend
    plt.legend(title='Legend')
    plt.title('2D Plot of BMI vs. Glucose by Cluster and Outcome')
    plt.xlabel('Glucose')
    plt.ylabel('BMI')
    # Save the plot as a jpg file
    plt.savefig('bmi_vs_glucose_plot_new.jpg', format='jpg')
    # Show the plot
    plt.show()


def visualization_new():
    # Load the data
    dt = pd.read_excel('diabetes_700/ways/way2/k3_train.xlsx')

    # Convert fraction strings to float
    for col in dt.columns:
        if dt[col].dtype == 'object':
            dt[col] = dt[col].str.replace('/', '').astype(float)

    # Set the style of the plot
    sns.set(style='whitegrid')

    # Define custom color mapping for each combination of Outcome and Cluster
    color_mapping = {
        (0, 0): 'green',
        (0, 1): 'blue',
        (0, 2): 'purple',
        (1, 0): 'red',
        (1, 1): 'pink',
        (1, 2): 'yellow'
    }

    # Plot the data
    plt.figure(figsize=(10, 6))
    for idx, row in dt.iterrows():
        outcome = row['Outcome']
        cluster = row['Cluster']
        color = color_mapping.get((outcome, cluster), 'gray')
        marker = 'o' if outcome == 0 else '^'  # Use circles for Outcome 0 and triangles for Outcome 1
        plt.scatter(row['Glucose'], row['BMI'], marker=marker, color=color)

    # Create custom legend
    legend_labels = {
        'Non-Diabetic Cluster 1 (Circle)': 'green',
        'Non-Diabetic Cluster 2 (Circle)': 'blue',
        'Non-Diabetic Cluster 3 (Circle)': 'purple',
        'Diabetic Cluster 1 (Triangle)': 'red',
        'Diabetic Cluster 2 (Triangle)': 'pink',
        'Diabetic Cluster 3 (Triangle)': 'yellow'
    }

    # Plot invisible points for creating custom legend
    for label, color in legend_labels.items():
        if 'Circle' in label:
            plt.scatter([], [], color=color, marker='o', label=label)
        else:
            plt.scatter([], [], color=color, marker='^', label=label)

    # Add legend
    plt.legend(title='Legend')
    plt.title('2D Plot of BMI vs. Glucose by Cluster and Outcome')
    plt.xlabel('Glucose')
    plt.ylabel('BMI')

    # Save the plot as a jpg file
    plt.savefig('bmi_vs_glucose_plot_new.jpg', format='jpg')

    # Show the plot
    plt.show()

def data_contains_null(data_new):
    # Check for null values
    null_values = data_new.isnull().sum()

    # Print the result
    print("Null values in each column:")
    print(null_values)

    # Check if there are any null values in the entire dataset
    if data_new.isnull().values.any():
        print("\nThe dataset contains null values.")
    else:
        print("\nThe dataset does not contain any null values.")


def normalize_breast_cancer(data_new):
    # Remove the 'id' column
    data_new = data_new.drop(columns=['id'])
    # Encode 'diagnosis' column (M -> 1, B -> 0)
    data_new['diagnosis'] = data_new['diagnosis'].map({'M': 1, 'B': 0})
    # Normalize the remaining features
    features = data_new.columns.drop('diagnosis')
    data_new[features] = (data_new[features] - data_new[features].min()) / (
                data_new[features].max() - data_new[features].min())

    # Save the processed dataset to a new file
    output_file = "breast_cancer/breast_cancer_normalized.xlsx"
    data_new.to_excel(output_file, index=False)
    print(f"Normalized dataset saved to {output_file}")


def rf():
    # Load the training and testing datasets
    train_file = "breast_cancer/1/train.xlsx"
    test_file = "breast_cancer/1/test_copy.xlsx"

    train_data = pd.read_excel(train_file)
    test_data = pd.read_excel(test_file)

    # Separate features and target variable for training
    X_train = train_data.drop(columns=['diagnosis'])
    y_train = train_data['diagnosis']

    # Separate features and target variable for testing
    X_test = test_data.drop(columns=['diagnosis'])
    y_test = test_data['diagnosis']

    # Train the Random Forest model
    # rf_model = RandomForestClassifier(random_state=42)
    # rf_model.fit(X_train, y_train)
    rf_model = RandomForestClassifier(max_depth=5, n_estimators=50, random_state=42)
    rf_model.fit(X_train, y_train)

    # Make predictions on the test set
    y_pred = rf_model.predict(X_test)

    # Evaluate the model
    accuracy = accuracy_score(y_test, y_pred)
    print(f"Accuracy of the Random Forest model: {accuracy}")


def main():
    warnings.filterwarnings('ignore')
    # wb = xl.load_workbook('diabetes_700/diabetes_700.xlsx')
    # sheet = wb['Sheet1']
    # df = pd.read_excel('diabetes_700/diabetes700_copy.xlsx', sheet_name='Sheet1')
    # data = pd.read_excel('diabetes_700/diabetes_700.xlsx', sheet_name='Sheet1')
    # wb_normal = xl.load_workbook('diabetes_700/ways/thesis_ch4.2/all.xlsx')
    # sheet_normal = wb_normal['Sheet1']
    # Load the Excel file
    file_path = "breast_cancer/breast_cancer.xlsx"
    data_new = pd.read_excel(file_path)
    wb = xl.load_workbook('breast_cancer/breast_cancer_normalized.xlsx')
    sheet = wb['Sheet1']
    """
        data has 569 rows
        data has 357 zero and 212 one outcome.
        30% of zeroes: 107
        30% of ones: 64
        70% of zeroes: 250
        70% of ones: 148
        
        357 zeroes and 212 ones were placed.
        test and train split.
        make excel for k.
        gmm done.
        method done.
        rf done.
    """
    # split_train_test(sheet)
    # GMM()
    # method()
    # randomforest()
    # data_confusion_matrix()
    # visualization_new()


if __name__ == "__main__":
    main()
